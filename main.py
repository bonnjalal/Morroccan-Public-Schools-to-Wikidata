from openpyxl import load_workbook
import re

wb = load_workbook("etab_pub_primaire_juin-2023.xlsx")

ws = wb.active

def cleanName(name:str):
    # Replace all occurrences of parentheses and content within with an empty string
    cleaned_name = re.sub(r"\(.*?\)", "", name)
    return cleaned_name.strip()

def cleanProvince(name:str):
    # Replace all occurrences of parentheses and content within with an empty string
    cleaned_name = re.sub(r".*?\:", "", name)
    return cleanProvince2(cleaned_name)

def cleanProvince2(name:str):
    cleaned = name.replace("Préf. d’Arr.", "")
    return cleaned.strip()


for row in range(1,7827):
    
    commune = ws.cell(row = row, column = 5).value
    newCommune = cleanName(commune)
    ws.cell(row = row, column = 5).value = newCommune

    province = ws.cell(row = row, column = 6).value
    newProvince = cleanProvince(province)
    ws.cell(row = row, column = 6).value = newProvince




wb.save("etab_pub_primaire_juin-2023-cleaned.xlsx")
